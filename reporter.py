"""Report generator for sitemap page-weight audit.

Aggregates findings across pages, determines scope (site-wide vs.
template-wide vs. page-specific), and outputs results as both a
rich terminal table and a JSON file.
"""

import json
import os
from collections import defaultdict
from typing import Optional

from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

from analyzer import Finding, PageAnalysis


def _format_bytes(size_bytes: int) -> str:
    """Format byte count as a human-readable string.

    Args:
        size_bytes: Number of bytes.

    Returns:
        A string like "1.5 KB" or "3.2 MB".
    """
    if size_bytes >= 1_048_576:
        return f"{size_bytes / 1_048_576:.1f} MB"
    if size_bytes >= 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes} B"


def _fingerprint(finding: Finding) -> str:
    """Create a fingerprint for grouping similar findings across pages.

    Findings are considered the same if they share the same element
    type and identifier. This allows us to detect shared elements.

    Args:
        finding: A Finding object.

    Returns:
        A string fingerprint for grouping.
    """
    return f"{finding.element_type}::{finding.element_identifier}"


def aggregate_findings(
    analyses: dict[str, list[PageAnalysis]],
) -> list[Finding]:
    """Aggregate findings across all pages and set scope.

    Merges findings with the same fingerprint, combining their
    pages_found_on lists. Determines scope:
    - "site-wide": found on pages across ALL templates.
    - "template-wide (<name>)": found on all pages within a template.
    - "page-specific": found on a single page only.
    - "multi-page": found on some but not all pages in a template.

    Args:
        analyses: Dict mapping template key -> list of PageAnalysis.

    Returns:
        A list of aggregated and deduplicated Finding objects,
        sorted by total size descending.
    """
    # Group findings by fingerprint
    grouped: dict[str, dict] = defaultdict(lambda: {
        "finding": None,
        "pages": set(),
        "templates": set(),
        "total_size": 0,
        "count": 0,
    })

    all_urls = set()
    template_urls: dict[str, set[str]] = defaultdict(set)

    for template, page_analyses in analyses.items():
        for analysis in page_analyses:
            all_urls.add(analysis.url)
            template_urls[template].add(analysis.url)

            for finding in analysis.findings:
                fp = _fingerprint(finding)
                entry = grouped[fp]

                if entry["finding"] is None:
                    entry["finding"] = Finding(
                        element_type=finding.element_type,
                        element_identifier=finding.element_identifier,
                        description=finding.description,
                        visibility=finding.visibility,
                        size_bytes=finding.size_bytes,
                        percent_of_page=finding.percent_of_page,
                        priority=finding.priority,
                        searchable_snippet=finding.searchable_snippet,
                    )

                entry["pages"].add(analysis.url)
                entry["templates"].add(template)
                entry["total_size"] = max(
                    entry["total_size"], finding.size_bytes
                )
                entry["count"] += 1

    # Determine scope for each grouped finding
    aggregated = []
    for fp, entry in grouped.items():
        finding = entry["finding"]
        pages = entry["pages"]
        templates = entry["templates"]

        finding.pages_found_on = sorted(pages)
        finding.size_bytes = entry["total_size"]

        # Determine scope
        if len(pages) == len(all_urls) and len(all_urls) > 1:
            finding.scope = "site-wide"
        elif len(templates) == 1:
            template_name = next(iter(templates))
            tmpl_urls = template_urls[template_name]
            if len(pages) == len(tmpl_urls) and len(tmpl_urls) > 1:
                finding.scope = f"template-wide ({template_name})"
            elif len(pages) > 1:
                finding.scope = f"multi-page ({template_name})"
            else:
                finding.scope = "page-specific"
        elif len(pages) > 1:
            finding.scope = "multi-page (cross-template)"
        else:
            finding.scope = "page-specific"

        aggregated.append(finding)

    # Sort: primary first, then by size descending
    aggregated.sort(
        key=lambda f: (
            0 if f.priority == "primary" else 1,
            -f.size_bytes,
        )
    )

    return aggregated


def print_page_summary(
    analyses: dict[str, list[PageAnalysis]],
    console: Optional[Console] = None,
) -> None:
    """Print a summary table of page sizes.

    Args:
        analyses: Dict mapping template key -> list of PageAnalysis.
        console: Optional rich Console. Creates one if not provided.
    """
    if console is None:
        console = Console()

    table = Table(
        title="📄 Page Size Summary",
        show_header=True,
        header_style="bold magenta",
    )
    table.add_column("Template", style="cyan", no_wrap=True)
    table.add_column("URL", style="dim")
    table.add_column("HTML Size", justify="right", style="bold")
    table.add_column("Flagged", justify="right")
    table.add_column("% Flagged", justify="right")

    for template, page_analyses in sorted(analyses.items()):
        for analysis in page_analyses:
            size_str = _format_bytes(analysis.total_html_bytes)
            flagged_str = _format_bytes(analysis.total_flagged_bytes)

            # Color code by size
            is_over_2mb = analysis.total_html_bytes > 2_097_152
            size_style = "bold red" if is_over_2mb else "green"

            table.add_row(
                template,
                analysis.url,
                Text(size_str, style=size_style),
                flagged_str,
                f"{analysis.flagged_percent:.1f}%",
            )

    console.print()
    console.print(table)


def print_findings_report(
    aggregated_findings: list[Finding],
    console: Optional[Console] = None,
    show_secondary: bool = True,
) -> None:
    """Print the main findings report as a rich table.

    Args:
        aggregated_findings: List of aggregated Finding objects.
        console: Optional rich Console.
        show_secondary: Whether to include secondary (external) findings.
    """
    if console is None:
        console = Console()

    # Separate primary and secondary
    primary = [f for f in aggregated_findings if f.priority == "primary"]
    secondary = [f for f in aggregated_findings if f.priority == "secondary"]

    # --- Primary findings table ---
    console.print()
    console.print(Panel(
        "[bold red]🔴 PRIMARY FINDINGS — Inline HTML Weight[/]\n"
        "[dim]These elements are embedded directly in the HTML file "
        "and contribute to its total size.[/]",
        border_style="red",
    ))

    if primary:
        table = Table(
            show_header=True,
            header_style="bold red",
            show_lines=True,
            expand=True,
        )
        table.add_column("Element", style="bold", max_width=40)
        table.add_column("Purpose", max_width=35)
        table.add_column("Visible?", justify="center", max_width=10)
        table.add_column("Size", justify="right", style="bold", max_width=10)
        table.add_column("% of Page", justify="right", max_width=8)
        table.add_column("Scope", style="cyan", max_width=20)
        table.add_column("Pages", max_width=40)

        for finding in primary:
            vis_icon = (
                "👁️ Yes" if finding.visibility == "user-visible"
                else "⚙️ No"
            )
            pages_str = "\n".join(finding.pages_found_on)

            table.add_row(
                finding.element_identifier,
                finding.description,
                vis_icon,
                _format_bytes(finding.size_bytes),
                f"{finding.percent_of_page:.1f}%",
                finding.scope,
                pages_str,
            )

        console.print(table)
    else:
        console.print("[green]No significant primary findings.[/]")

    # --- Secondary findings table ---
    if show_secondary and secondary:
        console.print()
        console.print(Panel(
            "[bold yellow]🟡 SECONDARY FINDINGS — External Resources[/]\n"
            "[dim]These are references to external files. They add "
            "minimal bytes to the HTML itself but trigger additional "
            "HTTP requests.[/]",
            border_style="yellow",
        ))

        table = Table(
            show_header=True,
            header_style="bold yellow",
            show_lines=True,
            expand=True,
        )
        table.add_column("Element", style="bold", max_width=45)
        table.add_column("Purpose", max_width=30)
        table.add_column("Visible?", justify="center", max_width=10)
        table.add_column("Scope", style="cyan", max_width=20)
        table.add_column("Pages", max_width=40)

        for finding in secondary:
            vis_icon = (
                "👁️ Yes" if finding.visibility == "user-visible"
                else "⚙️ No"
            )
            pages_str = "\n".join(finding.pages_found_on)

            table.add_row(
                finding.element_identifier,
                finding.description,
                vis_icon,
                finding.scope,
                pages_str,
            )

        console.print(table)


def write_json_report(
    analyses: dict[str, list[PageAnalysis]],
    aggregated_findings: list[Finding],
    output_path: str,
) -> None:
    """Write the full report as a JSON file.

    Args:
        analyses: Dict mapping template key -> list of PageAnalysis.
        aggregated_findings: List of aggregated Finding objects.
        output_path: Path for the JSON output file.
    """
    report = {
        "summary": {
            "total_pages_analyzed": sum(
                len(pa) for pa in analyses.values()
            ),
            "templates_found": list(analyses.keys()),
        },
        "pages": {},
        "aggregated_findings": {
            "primary": [
                f.to_dict()
                for f in aggregated_findings
                if f.priority == "primary"
            ],
            "secondary": [
                f.to_dict()
                for f in aggregated_findings
                if f.priority == "secondary"
            ],
        },
    }

    for template, page_analyses in analyses.items():
        report["pages"][template] = [
            pa.to_dict() for pa in page_analyses
        ]

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, ensure_ascii=False)


def write_excel_report(
    analyses: dict[str, list[PageAnalysis]],
    aggregated_findings: list[Finding],
    output_path: str,
) -> None:
    """Write the full report as an Excel file.

    Args:
        analyses: Dict mapping template key -> list of PageAnalysis.
        aggregated_findings: List of aggregated Finding objects.
        output_path: Path for the Excel output file.
    """
    import pandas as pd
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def _clean_str(val):
        if isinstance(val, str):
            return ILLEGAL_CHARACTERS_RE.sub("", val)
        return val

    report = {
        "summary": {
            "total_pages_analyzed": sum(
                len(pa) for pa in analyses.values()
            ),
            "templates_found": list(analyses.keys()),
        },
        "pages": {},
        "aggregated_findings": {
            "primary": [
                f.to_dict()
                for f in aggregated_findings
                if f.priority == "primary"
            ],
            "secondary": [
                f.to_dict()
                for f in aggregated_findings
                if f.priority == "secondary"
            ],
        },
    }

    for template, page_analyses in analyses.items():
        report["pages"][template] = [
            pa.to_dict() for pa in page_analyses
        ]

    # --- TAB 1: Process Page Findings ---
    page_rows = []
    js_payload_rows = []

    if "pages" in report:
        for template_name, pages_list in report["pages"].items():
            for page in pages_list:
                page_info = {
                    "Template": template_name,
                    "URL": page.get("url"),
                    "Total HTML Bytes": page.get("total_html_bytes"),
                    "Total HTML KB": page.get("total_html_kb"),
                    "Total Flagged Bytes": page.get("total_flagged_bytes"),
                    "Flagged Percent": page.get("flagged_percent"),
                    "Findings Count": page.get("findings_count")
                }
                
                findings = page.get("findings", [])
                if not findings:
                    page_rows.append(page_info)
                else:
                    for finding in findings:
                        row = page_info.copy()
                        row.update({
                            "Element Type": _clean_str(finding.get("element_type")),
                            "Element Identifier": _clean_str(finding.get("element_identifier")),
                            "Description": _clean_str(finding.get("description")),
                            "Visibility": _clean_str(finding.get("visibility")),
                            "Size (Bytes)": finding.get("size_bytes"),
                            "Percent of Page": finding.get("percent_of_page"),
                            "Priority": _clean_str(finding.get("priority")),
                            "Scope": _clean_str(finding.get("scope")),
                            "Pages Found On": _clean_str(", ".join(finding.get("pages_found_on", []))),
                            "Snippet": _clean_str(finding.get("searchable_snippet"))
                        })
                        page_rows.append(row)

                        # Capture json-nodes for the Large JS Payloads tab
                        if finding.get("element_type") == "json-node":
                            js_payload_rows.append({
                                "URL": page.get("url"),
                                "Template": template_name,
                                "Element Identifier": _clean_str(finding.get("element_identifier")),
                                "Description": _clean_str(finding.get("description")),
                                "Size (Bytes)": finding.get("size_bytes"),
                                "Percent of Page": finding.get("percent_of_page"),
                                "Snippet": _clean_str(finding.get("searchable_snippet"))
                            })

    df_pages = pd.DataFrame(page_rows)
    df_js_payloads = pd.DataFrame(js_payload_rows)

    # --- TAB 2: Process Aggregated Findings ---
    agg_rows = []

    if "aggregated_findings" in report:
        for category, findings_list in report["aggregated_findings"].items():
            for finding in findings_list:
                row = {
                    "Category": _clean_str(category.capitalize()),
                    "Element Type": _clean_str(finding.get("element_type")),
                    "Element Identifier": _clean_str(finding.get("element_identifier")),
                    "Description": _clean_str(finding.get("description")),
                    "Visibility": _clean_str(finding.get("visibility")),
                    "Size (Bytes)": finding.get("size_bytes"),
                    "Percent of Page": finding.get("percent_of_page"),
                    "Priority": _clean_str(finding.get("priority")),
                    "Scope": _clean_str(finding.get("scope")),
                    "Pages Found On": _clean_str(", ".join(finding.get("pages_found_on", []))),
                    "Snippet": _clean_str(finding.get("searchable_snippet"))
                }
                agg_rows.append(row)

    df_agg = pd.DataFrame(agg_rows)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_pages.to_excel(writer, sheet_name='Page Findings', index=False)
        df_agg.to_excel(writer, sheet_name='Aggregated Findings', index=False)
        if not df_js_payloads.empty:
            df_js_payloads.to_excel(writer, sheet_name='Large JS Payloads', index=False)

        # Apply rich styling to make it user-friendly
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Freeze the header row and add a filter dropdown to it
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions

            for col_idx, column in enumerate(worksheet.columns, 1):
                col_letter = get_column_letter(col_idx)
                header_cell = column[0]
                header_val = str(header_cell.value) if header_cell.value else ""

                # Style header
                header_cell.font = header_font
                header_cell.fill = header_fill

                # Determine optimal column width
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Cap the maximum width to keep things readable, minimum width 10
                adjusted_width = max(min(max_length + 2, 50), 10)
                worksheet.column_dimensions[col_letter].width = adjusted_width

                # Formatting data cells based on column name
                for cell in column[1:]:
                    if "Bytes" in header_val:
                        cell.number_format = '#,##0'  # Comma separated
                    
                    if "Percent" in header_val:
                        cell.number_format = '0.0'    # 1 decimal place
                        
                    # Enable word wrapping for wordy columns
                    if header_val in (
                        "Snippet", "Element Identifier", "Pages Found On", "URL", "Description"
                    ):
                        cell.alignment = wrap_alignment


def print_scope_summary(
    aggregated_findings: list[Finding],
    console: Optional[Console] = None,
) -> None:
    """Print a high-level summary of site-wide and template-wide findings.

    Args:
        aggregated_findings: List of aggregated Finding objects.
        console: Optional rich Console.
    """
    if console is None:
        console = Console()

    site_wide = [
        f for f in aggregated_findings
        if f.scope == "site-wide" and f.priority == "primary"
    ]
    template_wide = [
        f for f in aggregated_findings
        if "template-wide" in f.scope and f.priority == "primary"
    ]

    if site_wide:
        total_site_bytes = sum(f.size_bytes for f in site_wide)
        console.print()
        console.print(Panel(
            f"[bold]🌐 SITE-WIDE heavy elements "
            f"(found on ALL pages):[/]\n"
            f"[red]{len(site_wide)} elements totaling "
            f"{_format_bytes(total_site_bytes)}[/]\n\n"
            + "\n".join(
                f"  • {f.element_identifier} — "
                f"{_format_bytes(f.size_bytes)} — {f.description}"
                for f in site_wide
            ),
            title="🔍 Shared Element Analysis",
            border_style="bold blue",
        ))

    if template_wide:
        total_tmpl_bytes = sum(f.size_bytes for f in template_wide)
        console.print()
        console.print(Panel(
            f"[bold]📁 TEMPLATE-WIDE heavy elements "
            f"(shared within a template):[/]\n"
            f"[yellow]{len(template_wide)} elements totaling "
            f"{_format_bytes(total_tmpl_bytes)}[/]\n\n"
            + "\n".join(
                f"  • [{f.scope}] {f.element_identifier} — "
                f"{_format_bytes(f.size_bytes)} — {f.description}"
                for f in template_wide
            ),
            title="🔍 Shared Element Analysis",
            border_style="bold blue",
        ))

    if not site_wide and not template_wide:
        console.print()
        console.print(
            "[green]No shared heavy elements detected across "
            "pages or templates.[/]"
        )
